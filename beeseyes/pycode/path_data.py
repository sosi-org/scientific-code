import numpy as np
import os
from openpyxl import load_workbook

def load_xls(filename_xls, columnset):
    wb = load_workbook(filename_xls, read_only=True)
    #print (wb.sheetnames)
    #ws = wb.get_sheet_by_name('Sheet1')
    ws = wb.active

    ret = {}
    #use_col = 0
    for key in columnset:
        # key =- new header
        #columns: tuple = columnset[key]
        columns = columnset[key]

        print('columns', columns)

        res = []
        #for use_col in columns:
        for use_col in columns:
            print('column', use_col, 'of', key)
            #rows = ws.iter_rows()
            #header = ws.iter_rows().value
            for row in ws.iter_rows(min_row=1):
                header = row[use_col].value
                continue
            x2 = np.array([ float(r[use_col].value) for r in ws.iter_rows(min_row=2)])
            #x2 = np.array([float(r[use_col].value) for r in ws.iter_rows(*)])

            header = [r[use_col].value for r in ws.iter_rows(min_row=1)][0]

            #return x2[1:], x2[0]
            #res.append((x2, header))
            print('header',header,'for column', use_col);
            res.append(x2)
            print('', flush=True)
        res2d = np.array(res).T
        assert len(res2d.shape) == 2
        assert res2d.shape[1] == len(columns)
        print('res2d.shape', res2d.shape)
        ret[key] = res2d

    return ret

def load_trajectory_data(filename_xls):
    cols = {
     'fTime': (0,),
     'RWSmoothed': (6,7,8), # RWxSmoothed
     'direction': (10,11,12),
    }
    #x2, label = load_xls(cols)[0]
    #print('header:', label)
    allv = load_xls(filename_xls, cols)
    print('allv:', allv)
    print(allv)
    return allv
    '''
    bee_traj['RWSmoothed'] bounds:

    (7.079130000000021, 573.19317)
    (42.99566999999999, 371.98803)
    (-21.25277299999999, 272.127343)
    '''

"""
  returns attributes of of bee trajectory as numpy arrays:
  *  e.g. 'direction', 'RWSmoothed', 'fTime'
"""
def load_trajectory_cached(filename):
    CACHE_FILE = './traj-cache'  # traj.cache.npz
    cachefile_fullname = CACHE_FILE + '.npz'
    if (os.path.exists(cachefile_fullname)):
       print('Cache file found. Ignoring the .xslx file. To refresh the cache, delete ' + cachefile_fullname)
       bee_traj = np.load(cachefile_fullname)
       #print('>>bee_traj', bee_traj) # numpy.lib.npyio.NpzFile object

       #bee_traj.fields = {}
       #bee_traj.fields.A = 'A'
       bee_traj._RWSmoothed = 'RWSmoothed'
       bee_traj._direction = 'direction'
       bee_traj._fTime = 'fTime'

       #print('2',sorted(bee_traj.files))
       #print('3',bee_traj.files)
       print("bee_traj['RWSmoothed']")
       print(bee_traj['RWSmoothed'])
       print('---')
       return bee_traj

    else:
       print('Cache file not found. Creating it.')
       bee_traj = load_trajectory_data(filename)
       np.savez(CACHE_FILE, **bee_traj)
       # savez_compressed versus savez
       # 275 KB  275540 19 May 16:42 cache-traj.npz
       # 157 KB   157831 19 May 16:43 cache-traj.npz
       print('Saved trajectpories in a cache file. Now that it is cached, exiting. Run again to load the trajectory cache file')
       exit()


if __name__ == "__main__":
    CURRENT_PATH = '/Users/a9858770/cs/scientific-code/beeseyes'
    POSITIONS_XLS = CURRENT_PATH + '/Setup/beepath.xlsx'

    load_trajectory_data(POSITIONS_XLS)
